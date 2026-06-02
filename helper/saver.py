import csv
import json
from typing import Dict

import openpyxl


# transform in class
class FileWriter:
    """
    Handles file writing operations.
    """

    def __init__(self) -> None:
        super().__init__()

    def write_to_text_file(
        self,
        password: str
    ) -> None:
        with open(
            file="yours_password.txt",
            mode="w",
            encoding="utf-8"
        ) as file:
            file.write(password)

    def write_to_json_file(
        self,
        data: Dict[str, str]
    ) -> None:
        with open(
            file="yours_password.json",
            mode="w",
            encoding="utf-8"
        ) as file:
            json.dump(
                data,
                file,
                indent=4
            )

    def write_to_csv_file(
        self,
        date_time: str,
        password: str
    ) -> None:
        with open(
            file="yours_password.csv",
            mode="w",
            newline="",
            encoding="utf-8"
        ) as file:
            writer = csv.writer(file)

            writer.writerows(
                [
                    (
                        "DATE AND TIME",
                        "PASSWORD"
                    ),
                    [
                        date_time,
                        password
                    ]
                ]
            )

    def write_to_excel_file(
        self,
        data: Dict[str, str]
    ) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet["A1"] = "DATE AND TIME"
        sheet["B1"] = "PASSWORD"

        row: int = 2

        for key, value in data.items():
            sheet.cell(
                row=row,
                column=1
            ).value = key

            sheet.cell(
                row=row,
                column=2
            ).value = value

            row += 1

        workbook.save(
            "yours_password.xlsx"
        )